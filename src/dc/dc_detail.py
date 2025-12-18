import streamlit as st
from config.database import get_connection
from config.queries import DC_DETAIL, DC_ITEMS
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

def render_dc_detail():
    if 'selected_dc' not in st.session_state:
        st.error("No DC selected")
        return
    
    dc_number = st.session_state.selected_dc
    
    conn = get_connection()
    
    # Fetch DC details using centralized query
    dc = conn.execute(DC_DETAIL, (dc_number,)).fetchone()
    
    if not dc:
        st.error("DC not found")
        conn.close()
        return
    
    # Header with action buttons
    col1, col2, col3, col4 = st.columns([3, 1, 1, 1])
    with col1:
        st.markdown(f"## Delivery Challan: {dc_number}")
    with col2:
        if st.button("💾 Save DC", use_container_width=True):
            st.success("DC saved successfully!")
    with col3:
        if st.button("📄 Create GST Invoice", use_container_width=True):
            st.session_state.nav = 'GST Invoices'
            st.session_state.gst_action = 'create'
            st.rerun()
    with col4:
        if st.button("📥 Download DC", use_container_width=True, type="primary"):
            excel_file = generate_dc_excel(dc_number, conn)
            st.download_button(
                label="⬇️ Download Excel",
                data=excel_file,
                file_name=f"DC_{dc_number}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    if st.button("← Back to DC List"):
        st.session_state.dc_action = 'list'
        del st.session_state.selected_dc
        st.rerun()
    
    st.markdown("---")
    
    # DC Summary
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("**DC Date**")
        st.text(dc['dc_date'])
        
        st.markdown("**PO Number**")
        st.text(f"PO-{dc['po_number']}")
        
        st.markdown("**Supplier**")
        st.text(dc['supplier_name'] or "N/A")
    
    with col2:
        st.markdown("**Department No**")
        st.text(dc['department_no'] or "N/A")
        
        st.markdown("**Vehicle No**")
        st.text(dc['vehicle_no'] or "N/A")
        
        st.markdown("**LR Number**")
        st.text(dc['lr_no'] or "N/A")
    
    with col3:
        st.markdown("**Consignee**")
        st.text(dc['consignee_name'] or "N/A")
        
        st.markdown("**E-way Bill**")
        st.text(dc['eway_bill_no'] or "N/A")
        
        st.markdown("**Transporter**")
        st.text(dc['transporter'] or "N/A")
    
    st.markdown("---")
    
    # Items using centralized query
    st.markdown("### Dispatched Items")
    
    items = conn.execute(DC_ITEMS, (dc_number,)).fetchall()
    
    if items:
        # Table Header
        h1, h2, h3, h4, h5, h6 = st.columns([1, 2, 3, 1, 1, 2])
        h1.markdown("**S.No**")
        h2.markdown("**Material Code**")
        h3.markdown("**Description**")
        h4.markdown("**Unit**")
        h5.markdown("**Qty**")
        h6.markdown("**Value**")
        
        total_value = 0
        for idx, item in enumerate(items, 1):
            c1, c2, c3, c4, c5, c6 = st.columns([1, 2, 3, 1, 1, 2])
            c1.text(idx)
            c2.text(item['material_code'] or "N/A")
            c3.text(item['material_description'] or "N/A")
            c4.text(item['unit'] or "N/A")
            c5.text(f"{item['dispatch_qty']}")
            c6.text(f"₹{item['item_value']:,.2f}" if item['item_value'] else "₹0.00")
            
            total_value += item['item_value'] or 0
            
            st.markdown("<hr style='margin: 0.3rem 0; border-top: 1px solid #333333;'>", unsafe_allow_html=True)
        
        st.markdown(f"**Total Value: ₹{total_value:,.2f}**")
    else:
        st.info("No items in this DC")
    
    conn.close()


def generate_dc_excel(dc_number, conn):
    """Generate DC in Excel format matching the template"""
    
    # Fetch DC and items using centralized queries
    dc = conn.execute(DC_DETAIL, (dc_number,)).fetchone()
    items = conn.execute(DC_ITEMS, (dc_number,)).fetchall()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"DC-{dc_number}"
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    
    # Mock company data (as requested)
    phone = "Tel. No. 0755 - 4247748"
    gstin = "GSTIN: 23AACFS6810L1Z7"
    company_name = "SENSTOGRAPHIC"
    company_address = "Manufacturers & Suppliers of Fibre Glass Re-inforced Plastic Products\nPlot No. 20/21, 'H' Sector, Industrial Estate, Govindpura, Bhopal - 462023"
    
    # Header
    row = 1
    ws[f'A{row}'] = phone
    ws[f'C{row}'] = gstin
    ws[f'A{row}'].font = Font(size=10)
    ws[f'C{row}'].font = Font(size=10)
    
    row += 2
    ws[f'A{row}'] = company_name
    ws[f'A{row}'].font = Font(size=18, bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws[f'A{row}'] = company_address
    ws[f'A{row}'].font = Font(size=9)
    ws[f'A{row}'].alignment = Alignment(horizontal='center', wrap_text=True)
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 2
    ws[f'A{row}'] = "DELIVERY CHALLAN"
    ws[f'A{row}'].font = Font(size=16, bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws.merge_cells(f'A{row}:C{row}')
    
    # Left box - Delivery person and company
    row += 2
    delivery_person = "The Sr. Manager (CRX)"
    delivery_company = f"M/S {dc['consignee_name'] or 'Bharat Heavy Eletrical Ltd.'}"
    delivery_location = dc['consignee_address'] or "BHOPAL"
    
    ws[f'A{row}'] = delivery_person
    ws[f'A{row}'].font = Font(size=10)
    row += 1
    ws[f'A{row}'] = delivery_company
    ws[f'A{row}'].font = Font(size=10, bold=True)
    row += 1
    ws[f'A{row}'] = delivery_location
    ws[f'A{row}'].font = Font(size=10)
    
    # Right box - Challan details
    row -= 2
    ws[f'C{row}'] = f"Challan No. : {dc_number}"
    ws[f'C{row}'].font = Font(size=10)
    row += 1
    ws[f'C{row}'] = f"Date : {dc['dc_date']}"
    ws[f'C{row}'].font = Font(size=10)
    row += 1
    ws[f'C{row}'] = f"Our Ref : SSG-{dc_number[-4:]}"  # Mock internal reference
    ws[f'C{row}'].font = Font(size=10)
    
    # Order details
    row += 2
    ws[f'A{row}'] = f"Your Order No. : {dc['po_number']}"
    ws[f'B{row}'] = f"Date: {dc['po_date']}"
    ws[f'C{row}'] = "Amd. Date:"
    ws[f'A{row}'].font = Font(size=10)
    ws[f'B{row}'].font = Font(size=10)
    ws[f'C{row}'].font = Font(size=10)
    
    row += 1
    ws[f'A{row}'] = f"Goods Dispatched Delivered to: {dc['department_no'] or '204'}"
    ws[f'A{row}'].font = Font(size=10)
    
    row += 1
    ws[f'C{row}'] = f"GST Invoice No: {dc_number} Dt. {dc['dc_date']}"
    ws[f'C{row}'].font = Font(size=10)
    
    # Items table header
    row += 2
    ws[f'A{row}'] = "P.O.Sl.\nNo."
    ws[f'B{row}'] = "Description"
    ws[f'C{row}'] = "Quantity"
    
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(size=11, bold=True)
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'{col}{row}'].border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Items
    for item in items:
        row += 1
        ws[f'A{row}'] = item['po_item_no']
        ws[f'B{row}'] = item['material_description'] or item['material_code']
        ws[f'C{row}'] = f"{item['dispatch_qty']} {item['unit']}"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(size=10)
            ws[f'{col}{row}'].alignment = Alignment(vertical='top', wrap_text=True)
            ws[f'{col}{row}'].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        ws.row_dimensions[row].height = 40
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()
