import io
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.models import POHeader, POItem, PODelivery
from typing import List, Optional

COMPANY_INFO = {
    "name": "SENSTOGRAPHIC",
    "address_lines": ["H-20/21, INDUSTRIAL AREA", "GOVINDPURA, BHOPAL"],
    "gstin": "23AAECS4648Q1Z4"
}

class ExcelService:
    @staticmethod
    def _apply_border(cell, style='thin'):
        thin = Side(border_style=style, color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    @staticmethod
    def generate_po_excel(header: POHeader, items: List[POItem], deliveries: List[PODelivery]) -> io.BytesIO:
        """
        Generates PO Excel matching 6624221.html structure
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"PO_{header.po_number}"
        
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # 1. Header Block (Left: Logo/Company, Right: PO Details)
        # Using simple text layout matching observed HTML structure
        ws.merge_cells('A1:D4')
        ws['A1'] = f"{COMPANY_INFO['name']}\n{chr(10).join(COMPANY_INFO['address_lines'])}"
        ws['A1'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        ws.merge_cells('E1:H4')
        ws['E1'] = f"PURCHASE ORDER\nPO No: {header.po_number}\nDate: {header.po_date}\nVersion: 1"
        ws['E1'].alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
        ws['E1'].font = bold_font

        # 2. Supplier & Ship To Info
        ws.merge_cells('A6:D8')
        ws['A6'] = f"To (Supplier):\n{header.supplier_name}\n{header.supplier_code}\nGSTIN: {header.supplier_gstin}"
        ws['A6'].alignment = left_align
        
        ws.merge_cells('E6:H8')
        ship_to_address = chr(10).join(COMPANY_INFO['address_lines'])
        ws['E6'] = f"Ship To:\n{COMPANY_INFO['name']}\n{ship_to_address}"
        ws['E6'].alignment = left_align

        # 3. Items Table
        # Columns: Sl No, Material Code, Description, Due Date, Qty, Unit, Rate, Amount
        headers = ["Sl No", "Material Code", "Description", "Due Date", "Qty", "Unit", "Rate", "Amount"]
        row_num = 10
        for col, title in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=col, value=title)
            c.font = bold_font
            c.border = thin_border
            c.alignment = center_align
            
        row_num += 1
        total_amt = 0
        for idx, item in enumerate(items, 1):
            qty = float(item.ordered_quantity or 0)
            rate = float(item.po_rate or 0)
            amt = qty * rate
            total_amt += amt
            
            dely_date = deliveries[0].dely_date if deliveries else ""

            row_data = [
                item.po_item_no,
                item.material_code,
                item.material_description,
                dely_date,
                qty,
                item.unit,
                rate,
                amt
            ]
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=row_num, column=col, value=val)
                c.border = thin_border
                c.alignment = center_align if col != 3 else left_align # Align desc left
                if isinstance(val, (int, float)) and col > 6:
                    c.number_format = '0.00'
            row_num += 1
            
        # Total
        ws.merge_cells(f'A{row_num}:G{row_num}')
        ws[f'A{row_num}'] = "Total Amount"
        ws[f'A{row_num}'].alignment = right_align
        ws[f'A{row_num}'].font = bold_font
        
        ws.cell(row=row_num, column=8, value=total_amt).number_format = '0.00'
        ws.cell(row=row_num, column=8).font = bold_font
        ws.cell(row=row_num, column=8).border = thin_border
        
        # 4. Terms & Conditions
        row_num += 2
        ws['A' + str(row_num)] = "Terms and Conditions:"
        ws['A' + str(row_num)].font = bold_font
        
        row_num += 1
        ws['A' + str(row_num)] = "1. Please supply material as per specification.\n2. Payment within 30 days."
        
        # Column Widths
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 12
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_dc_excel(dc_data: dict, items: List[dict]) -> io.BytesIO:
        """
        Generates DC Excel matching DC_template.xls
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Delivery Challan"
        
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = "DELIVERY CHALLAN"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align
        
        # Org Info
        ws.merge_cells('A2:F3')
        ws['A2'] = f"{COMPANY_INFO['name']}\n{COMPANY_INFO['address_lines'][-1]}" # Using City Line
        ws['A2'].alignment = center_align
        
        # Addresses
        ws.merge_cells('A4:C6')
        ws['A4'] = f"To (Consignee):\n{dc_data.get('consignee_name')}\n{dc_data.get('delivery_address', '')}"
        ws['A4'].alignment = Alignment(vertical='top', wrap_text=True)
        
        ws.merge_cells('D4:F4')
        ws['D4'] = f"DC No: {dc_data.get('dc_number')}"
        ws.merge_cells('D5:F5')
        ws['D5'] = f"Date: {dc_data.get('dc_date')}"
        ws.merge_cells('D6:F6')
        ws['D6'] = f"Ref PO: {dc_data.get('po_number')}"

        # Table Header: Sl. No., Description, Part No, Qty, Remarks
        headers = ["Sl. No.", "Description", "Part No", "Qty", "Remarks"]
        row_num = 8
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
            
        # Items
        row_num += 1
        for idx, item in enumerate(items, 1):
            row_data = [
                idx,
                item.get('material_description', item.get('item_description')), # Handle both keys
                item.get('material_code', ''), 
                item.get('dispatch_qty'),
                "" # Remarks
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            row_num += 1
            
        # Footer
        row_num += 2
        ws.merge_cells(f'A{row_num}:C{row_num}')
        ws[f'A{row_num}'] = "Receiver's Signature"
        ws.merge_cells(f'D{row_num}:F{row_num}')
        ws[f'D{row_num}'] = f"For {COMPANY_INFO['name']}"
        ws[f'D{row_num}'].alignment = Alignment(horizontal='right')

        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_invoice_excel(invoice_data: dict, items: List[dict]) -> io.BytesIO:
        """
        Generates Invoice Excel matching GST_INV_template.xls
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Tax Invoice"
        
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 1. Title
        ws.merge_cells('A1:M1')
        ws['A1'] = "TAX INVOICE"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_align
        
        # 2. Company & Invoice Details Header
        ws.merge_cells('A2:F4')
        address_str = chr(10).join(COMPANY_INFO['address_lines'])
        ws['A2'] = f"{COMPANY_INFO['name']}\n{address_str}\nGSTIN: {invoice_data.get('company_gstin', COMPANY_INFO['gstin'])}"
        ws['A2'].alignment = left_align
        
        ws.merge_cells('G2:I2')
        ws['G2'] = "Invoice No:"
        ws.merge_cells('J2:M2')
        ws['J2'] = invoice_data.get('invoice_number')
        
        ws.merge_cells('G3:I3')
        ws['G3'] = "Invoice Date:"
        ws.merge_cells('J3:M3')
        ws['J3'] = invoice_data.get('invoice_date')
        
        ws.merge_cells('G4:I4')
        ws['G4'] = "Reverse Charge (Y/N):"
        ws.merge_cells('J4:M4')
        ws['J4'] = "N"

        # 3. Bill To / Ship To
        ws.merge_cells('A5:F9')
        ws['A5'] = f"Bill To:\n{invoice_data.get('bill_to_name') or invoice_data.get('customer_name')}\n{invoice_data.get('bill_to_address')}\nGSTIN: {invoice_data.get('customer_gstin')}"
        ws['A5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        ws.merge_cells('G5:M9')
        ws['G5'] = f"Ship To:\n{invoice_data.get('ship_to_name') or invoice_data.get('customer_name')}\n{invoice_data.get('ship_to_address')}"
        ws['G5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # 4. Table Headers matching GST_INV_template
        headers = [
            "S.No.", "Description of Goods", "HSN/SAC Code", "Qty", "Unit", 
            "Rate", "Amount", "Discount", "Taxable Value", "CGST", "SGST", "IGST", "Total"
        ]

        row_num = 11
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
            
        # 5. Items
        row_num += 1
        for idx, item in enumerate(items, 1):
            qty = float(item.get('quantity', 0))
            rate = float(item.get('rate', 0))
            amount = qty * rate
            discount = 0.0 
            taxable = amount - discount
            
            cgst_val = float(item.get('cgst_amount', 0))
            sgst_val = float(item.get('sgst_amount', 0))
            igst_val = float(item.get('igst_amount', 0))
            total = taxable + cgst_val + sgst_val + igst_val

            row_data = [
                idx,
                item.get('description', item.get('material_description')), # Handle both keys
                item.get('hsn_code', '8525'),
                qty,
                "Nos",
                rate,
                amount,
                discount,
                taxable,
                cgst_val,
                sgst_val,
                igst_val,
                total
            ]
            
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.alignment = center_align
                cell.border = thin_border
                if isinstance(val, (int, float)) and col > 5:
                     cell.number_format = '0.00'
            row_num += 1

        # Final Totals Row
        ws.merge_cells(f'A{row_num}:H{row_num}')
        ws[f'A{row_num}'] = "Total"
        ws[f'A{row_num}'].alignment = Alignment(horizontal='right')
        ws[f'A{row_num}'].font = bold_font
        ws[f'A{row_num}'].border = thin_border
        
        # Sums
        for col_idx in range(9, 14): # Taxable to Total
            col_letter = ws.cell(row=row_num, column=col_idx).column_letter
            ws[f'{col_letter}{row_num}'] = f"=SUM({col_letter}12:{col_letter}{row_num-1})"
            ws[f'{col_letter}{row_num}'].font = bold_font
            ws[f'{col_letter}{row_num}'].number_format = '0.00'
            ws[f'{col_letter}{row_num}'].border = thin_border

        # Column Widths
        ws.column_dimensions['B'].width = 35
        for c in ['C','D','E','F','G','H','I','J','K','L','M']:
            ws.column_dimensions[c].width = 12
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_date_summary_excel(entity: str, start_date: str, end_date: str, rows: List[dict]) -> io.BytesIO:
        """
        Generates Date-wise Summary Report.
        Supports 'invoice' (matches SALE SUMMARY format) and others.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"{entity.upper()} Summary"
        
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        if entity == "invoice":
            # --- SALE SUMMARY FORMAT ---
            # Determine Header Name (Use Party Name from first row if available, else usage default)
            header_name = COMPANY_INFO['name']
            if rows and (rows[0].get('party_name') or rows[0].get('supplier_name')):
                header_name = rows[0].get('party_name') or rows[0].get('supplier_name')

            ws.merge_cells('A1:O1')
            ws['A1'] = header_name
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = center_align
            
            ws.merge_cells('A2:O2')
            ws['A2'] = "SALE SUMMARY"
            ws['A2'].font = Font(bold=True, size=12, underline="single")
            ws['A2'].alignment = center_align

            headers = [
                "S. No.", "Inv No. & Dt", "DC No.", 
                "PO & Item Description", "Qty", "Inv Ass. Value", 
                "E.D. 10 %", "ED Cess 2 %", "ED Cess1 %", "Sub Total", 
                "VAT 13%", "CST 2 %", "Other Additons", "Total"
            ]
            
            ws.append([]) # Row 3 is headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = bold_font
                cell.alignment = center_align
                cell.border = thin_border
            
            # Widths
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['D'].width = 30
            
            # Rows
            row_idx = 4
            for i, row in enumerate(rows, 1):
                # Helper to get value
                def get(k): return row.get(k, '')
                
                # Format Invoice Date (assuming row has 'invoice_date' as 'YYYY-MM-DD')
                inv_date = get('invoice_date')
                try:
                    vals = inv_date.split('-')
                    if len(vals) == 3: inv_date = f"{vals[2]}/{vals[1]}/{vals[0]}"
                except: pass
                
                inv_desc = f"{get('invoice_number')}\n{inv_date}"
                
                # Description logic
                desc = get('description') or ''
                if get('po_number'):
                    desc += f"\n(PO: {get('po_number')})"

                data = [
                    i,
                    inv_desc,
                    get('linked_dc_numbers'),
                    desc,
                    float(get('quantity') or 0),
                    float(get('amount') or get('invoice_value') or 0), # Ass. Value
                    0, 0, 0, # ED stuff
                    float(get('amount') or get('invoice_value') or 0), # Sub Total
                    0, 0, 0, # Taxes
                    float(get('total_invoice_value') or get('invoice_value') or 0) # Total
                ]
                
                for col, val in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    if isinstance(val, (int, float)) and col > 5:
                        cell.number_format = '0.00'
                
                row_idx += 1
                
        else:
             # Generic Fallback for PO/Challan
            headers = [k.replace('_', ' ').title() for k in rows[0].keys()] if rows else []
            ws.append(headers)
            for cell in ws[1]:
                cell.font = bold_font
                cell.border = thin_border
            
            for row in rows:
                ws.append(list(row.values()))
                
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_summary_excel(start_date: date, end_date: date, data: List[dict]) -> io.BytesIO:
        """
        Generates Summary Report matching SUMMARY_March 2019 format
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary Report"
        
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Title
        ws.merge_cells('A1:J1')
        ws['A1'] = f"SUMMARY REPORT ({start_date} to {end_date})"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = center_align
        
        # Headers based on inferred structure from context
        row = 3
        headers = [
            "Date", "Inv No", "Challan No", "PO No", "Party Name", 
            "Item", "Qty", "Rate", "Basic Value", "Total Value"
        ]
        
        for col, title in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=title)
            c.font = bold_font
            c.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            ExcelService._apply_border(c)
            
        row += 1
        for item in data:
            ws.cell(row=row, column=1, value=item.get('date'))
            ws.cell(row=row, column=2, value=item.get('invoice_no'))
            ws.cell(row=row, column=3, value=item.get('challan_no'))
            ws.cell(row=row, column=4, value=item.get('po_no'))
            ws.cell(row=row, column=5, value=item.get('party_name'))
            ws.cell(row=row, column=6, value=item.get('item_desc'))
            ws.cell(row=row, column=7, value=item.get('qty'))
            ws.cell(row=row, column=8, value=item.get('rate'))
            ws.cell(row=row, column=9, value=item.get('basic_value'))
            ws.cell(row=row, column=10, value=item.get('total_value'))
            
            for col in range(1, 11):
                ExcelService._apply_border(ws.cell(row=row, column=col))
            row += 1

        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 30
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
